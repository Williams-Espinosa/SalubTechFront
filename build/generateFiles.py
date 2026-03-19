#!/usr/bin/env python3
# build/generateFiles.py — genera bitacora.xlsx y reporte_auditoria.pdf
# Uso: python3 build/generateFiles.py --out build/tmp

import argparse, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

parser = argparse.ArgumentParser()
parser.add_argument("--out", default="build/tmp")
args = parser.parse_args()
os.makedirs(args.out, exist_ok=True)

# ── DATA ─────────────────────────────────────────────────────────────────────
BIT_LOGS = [
    ("Login Exitoso",  "Admin HR (Elena)", "29/01/26 08:30:12", "SUCCESS"),
    ("Edición Perfil", "Admin HR (Elena)", "29/01/26 09:15:44", "SUCCESS"),
    ("Intento Fallido","HT-SERVICE-GATE",  "29/01/26 09:22:10", "BLOCKED"),
    ("Baja Empleado",  "System-Auto",      "29/01/26 10:05:00", "SUCCESS"),
    ("Exportar PDF",   "Admin HR (Elena)", "29/01/26 10:30:22", "SUCCESS"),
    ("Actualizar DB",  "SysAdmin",         "29/01/26 11:00:00", "SUCCESS"),
]
INCIDENTS = [
    ("Juan Pérez",       "I1","B-102","2026-01-25 14:30","Reacción alérgica leve a medicamento administrado.",            "Lic. Ricardo Gómez","LEVE"),
    ("María García",     "I2","B-205","2026-01-26 09:15","Caída accidental al intentar levantarse sin asistencia.",        "Mtra. Sofía Ruiz",  "MODERADO"),
    ("Roberto Hernández","I3","C-301","2026-01-27 22:10","Error en la actualización de signos vitales en el sistema.",     "Lic. Ricardo Gómez","LEVE"),
]
STAFF = [
    ("Dra. Elena Martínez","MARE904001HOFLRS01","SUPERVISOR","Cardiología",     "Matutino",  "Activo"),
    ("Lic. Ricardo Gómez", "GORR921215HOFLRS02","ENFERMERO", "Urgencias",       "Nocturno",  "Activo"),
    ("Dr. Julián Sossa",   "SOJJ800120HOFLRS03","MÉDICO",    "Pediatría",       "Vespertino","Inactivo"),
    ("Mtra. Sofía Ruiz",   "RUIS880808MOFLRS04","ENFERMERO", "Terapia Intensiva","Matutino", "Activo"),
]

BLUE="3D5BF5"; BLUE_L="EEF1FE"; GREEN="38A169"; RED="E53E3E"
GRAY="F0F2F5"; WHITE="FFFFFF"; DARK="0F1B3D"; MID="4A5578"

def hdr(cell, bg=BLUE, fg=WHITE):
    cell.font = Font(name="Arial", bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def thin_border(ws, r1, r2, c1, c2):
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="E2E5EF")
    b = Border(left=s, right=s, top=s, bottom=s)
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
        for cell in row: cell.border = b

# ── XLSX ─────────────────────────────────────────────────────────────────────
wb = Workbook()

def make_sheet(ws, title_text, headers, widths, rows, color_col=None, color_fn=None):
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 32
    nc = len(headers)
    ws.merge_cells(f"A1:{get_column_letter(nc)}1")
    ws["A1"] = title_text
    ws["A1"].font = Font(name="Arial", bold=True, color=WHITE, size=14)
    ws["A1"].fill = PatternFill("solid", start_color=BLUE)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A2:{get_column_letter(nc)}2")
    ws["A2"].fill = PatternFill("solid", start_color=BLUE_L)
    for i,(h,w) in enumerate(zip(headers,widths),1):
        hdr(ws.cell(row=3,column=i,value=h))
        ws.column_dimensions[get_column_letter(i)].width = w
    for r,row in enumerate(rows,4):
        ws.row_dimensions[r].height = 22
        bg = GRAY if r%2==0 else WHITE
        for c,val in enumerate(row,1):
            cell = ws.cell(row=r,column=c,value=val)
            fc = DARK
            if color_col and color_fn and c==color_col:
                fc = color_fn(row)
            cell.font = Font(name="Arial", size=10, color=fc, bold=(color_col and c==color_col))
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(vertical="center", horizontal="center" if c>=color_col-1 else "left", indent=1 if c==1 else 0)
    thin_border(ws, 3, 3+len(rows), 1, nc)

ws1 = wb.active; ws1.title = "Bitácora de Sistema"
make_sheet(ws1,"SALUDTECH — Bitácora de Sistema",
    ["Evento","Usuario / Módulo","Timestamp","Estado"],[28,26,22,14],
    BIT_LOGS, color_col=4, color_fn=lambda r: GREEN if r[3]=="SUCCESS" else RED)

ws2 = wb.create_sheet("Personal")
make_sheet(ws2,"SALUDTECH — Personal del Hospital",
    ["Nombre","ID / CURP","Rol","Especialidad","Turno","Estado"],[28,24,14,22,14,12],
    STAFF, color_col=6, color_fn=lambda r: GREEN if r[5]=="Activo" else "8C93AA")

ws3 = wb.create_sheet("Incidentes")
make_sheet(ws3,"SALUDTECH — Auditoría de Incidentes",
    ["Paciente","ID","Cama","Fecha/Hora","Descripción","Reportado por","Gravedad"],[22,8,10,20,42,24,14],
    INCIDENTS, color_col=7, color_fn=lambda r: {"LEVE":BLUE,"MODERADO":"D69E2E","GRAVE":RED}.get(r[6],DARK))

xlsx_path = os.path.join(args.out, "bitacora.xlsx")
wb.save(xlsx_path)
print(f"✅ XLSX: {xlsx_path}")

# ── PDF ──────────────────────────────────────────────────────────────────────
pdf_path = os.path.join(args.out, "reporte_auditoria.pdf")
doc = SimpleDocTemplate(pdf_path, pagesize=letter,
    leftMargin=.7*inch, rightMargin=.7*inch, topMargin=.7*inch, bottomMargin=.7*inch)

BLUE_C  = colors.HexColor("#3D5BF5")
BLUE_LC = colors.HexColor("#EEF1FE")
DARK_C  = colors.HexColor("#0F1B3D")
MID_C   = colors.HexColor("#4A5578")
SOFT_C  = colors.HexColor("#8C93AA")
GREEN_C = colors.HexColor("#38A169")
RED_C   = colors.HexColor("#E53E3E")
GRAY_C  = colors.HexColor("#F0F2F5")

ss = getSampleStyleSheet()
def sty(name, **kw): return ParagraphStyle(name, **kw)

story = []

# Header
hdr_data = [[
    Paragraph("<b><font color='#3D5BF5' size=18>SALUDTECH</font></b><br/><font color='#8C93AA' size=8>HOSPITAL MANAGEMENT SYSTEM</font>", ss["Normal"]),
    Paragraph("<font color='#8C93AA' size=8>REPORTE DE AUDITORÍA</font><br/><b><font color='#0F1B3D' size=10>Auditoría &amp; Desempeño</font></b><br/><font color='#8C93AA' size=8>29 de Enero, 2026</font>",
              sty("rh", fontName="Helvetica", fontSize=9, alignment=2))
]]
ht = Table(hdr_data, colWidths=[3.5*inch,3.5*inch])
ht.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),BLUE_LC),("TOPPADDING",(0,0),(-1,-1),14),("BOTTOMPADDING",(0,0),(-1,-1),14),("LEFTPADDING",(0,0),(0,-1),16),("RIGHTPADDING",(1,0),(1,-1),16)]))
story.append(ht); story.append(Spacer(1,16))

# KPIs
kpi = [[
    Paragraph("<font color='#8C93AA' size=7>COLABORADORES</font><br/><b><font color='#0F1B3D' size=20>142</font></b><br/><font color='#38A169' size=8>+4 esta semana</font>", ss["Normal"]),
    Paragraph("<font color='#8C93AA' size=7>PACIENTES ACTIVOS</font><br/><b><font color='#0F1B3D' size=20>86</font></b><br/><font color='#4A5578' size=8>-2% ocupación</font>", ss["Normal"]),
    Paragraph("<font color='#8C93AA' size=7>INCIDENTES HOY</font><br/><b><font color='#0F1B3D' size=20>03</font></b><br/><font color='#4A5578' size=8>Gravedad: Leve</font>", ss["Normal"]),
    Paragraph("<font color='#8C93AA' size=7>REPORTES PDF</font><br/><b><font color='#0F1B3D' size=20>28</font></b><br/><font color='#4A5578' size=8>Generados este mes</font>", ss["Normal"]),
]]
kt = Table(kpi, colWidths=[1.75*inch]*4)
kt.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.white),("GRID",(0,0),(-1,-1),.5,colors.HexColor("#E2E5EF")),("TOPPADDING",(0,0),(-1,-1),14),("BOTTOMPADDING",(0,0),(-1,-1),14),("LEFTPADDING",(0,0),(-1,-1),14)]))
story.append(kt); story.append(Spacer(1,20))

sec_sty = sty("sec", fontName="Helvetica-Bold", fontSize=13, textColor=DARK_C, spaceBefore=18, spaceAfter=10)

# Incidentes table
story.append(Paragraph("Incidentes Registrados", sec_sty))
ih = [["Paciente","Cama","Fecha/Hora","Descripción","Reportado por","Gravedad"]]
ir = [[p,c,dt,desc[:50]+("…" if len(desc)>50 else ""),rep,sev] for p,pid,c,dt,desc,rep,sev in INCIDENTS]
it = Table(ih+ir, colWidths=[1.2*inch,.6*inch,1.1*inch,2.0*inch,1.3*inch,.7*inch], repeatRows=1)
sev_map={"LEVE":BLUE_C,"MODERADO":colors.HexColor("#D69E2E"),"GRAVE":RED_C}
its=[("BACKGROUND",(0,0),(-1,0),BLUE_C),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),("ALIGN",(0,0),(-1,0),"CENTER"),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,GRAY_C]),("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),8),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#E2E5EF")),("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),("LEFTPADDING",(0,0),(-1,-1),8)]
for i,(_,_,_,_,_,_,sev) in enumerate(INCIDENTS,1):
    its+=[("TEXTCOLOR",(5,i),(5,i),sev_map.get(sev,DARK_C)),("FONTNAME",(5,i),(5,i),"Helvetica-Bold")]
it.setStyle(TableStyle(its)); story.append(it); story.append(Spacer(1,16))

# Personal table
story.append(Paragraph("Personal del Hospital", sec_sty))
ph=[["Nombre","ID / CURP","Rol","Especialidad","Turno","Estado"]]
pr=[[n,i,r,s,t,e] for n,i,r,s,t,e in STAFF]
pt=Table(ph+pr,colWidths=[1.5*inch,1.4*inch,.85*inch,1.2*inch,.85*inch,.7*inch],repeatRows=1)
pts=[("BACKGROUND",(0,0),(-1,0),BLUE_C),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),("ALIGN",(0,0),(-1,0),"CENTER"),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,GRAY_C]),("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),8),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#E2E5EF")),("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),("LEFTPADDING",(0,0),(-1,-1),8)]
for i,(_,_,_,_,_,estado) in enumerate(STAFF,1):
    pts+=[("TEXTCOLOR",(5,i),(5,i),GREEN_C if estado=="Activo" else SOFT_C),("FONTNAME",(5,i),(5,i),"Helvetica-Bold")]
pt.setStyle(TableStyle(pts)); story.append(pt); story.append(Spacer(1,16))

# Bitácora table
story.append(Paragraph("Bitácora de Sistema", sec_sty))
bh=[["Evento","Usuario / Módulo","Timestamp","Estado"]]
br=[[ev,usr,ts,st] for ev,usr,ts,st in BIT_LOGS]
bt=Table(bh+br,colWidths=[1.6*inch,1.6*inch,1.5*inch,.8*inch],repeatRows=1)
bts=[("BACKGROUND",(0,0),(-1,0),BLUE_C),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,0),8),("ALIGN",(0,0),(-1,0),"CENTER"),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,GRAY_C]),("FONTNAME",(0,1),(-1,-1),"Helvetica"),("FONTSIZE",(0,1),(-1,-1),8),("GRID",(0,0),(-1,-1),.4,colors.HexColor("#E2E5EF")),("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),("LEFTPADDING",(0,0),(-1,-1),8)]
for i,(_,_,_,st) in enumerate(BIT_LOGS,1):
    bts+=[("TEXTCOLOR",(3,i),(3,i),GREEN_C if st=="SUCCESS" else RED_C),("FONTNAME",(3,i),(3,i),"Helvetica-Bold")]
bt.setStyle(TableStyle(bts)); story.append(bt); story.append(Spacer(1,24))

story.append(HRFlowable(width="100%",thickness=1,color=colors.HexColor("#E2E5EF")))
story.append(Spacer(1,8))
story.append(Paragraph("<font color='#8C93AA' size=7>© 2026 HospTrack Systems · Seguridad de datos de grado médico · Documento generado automáticamente por SALUDTECH RH Module</font>",
    sty("ft",fontName="Helvetica",fontSize=7,textColor=SOFT_C,alignment=1)))

doc.build(story)
print(f"✅ PDF:  {pdf_path}")
